import openpyxl         #pip install openpyxl

'''
Upload data from MSC Nastran dhf5 schema to an excel file.
Place the excel file in the same folder as this script, or specify the full file path location below.
Run script. There are 3 txt files produced, as specified in the main method. Each can be turned off, just make sure the global variables are correct.
Paste code into a Visual Studio file for formatting corrections
'''



#Class to create objects that represent a dataset. Called "grouping" because it collates groups of rows in the spreadsheet. 
#Way to package info together without having to look for it repeatedly.
class Grouping():
    def __init__(self, ws, title='', rows=[]):
        self.__ws=ws
        self.__rows=rows
        self.__title=title
        self.__BCol=self.populateCols("B")
        self.__CCol=self.populateCols("C")
        self.__DCol=self.populateCols("D")
        self.__partners=[]
    
    def getTitle(self):
        return self.__title
    def getRows(self):
        return self.__rows
    def getBCol(self):
        return self.__BCol
    def getCCol(self):
        return self.__CCol
    def getDCol(self):
        return self.__DCol
    def getPartners(self):  #Unused
        return self.__partners
    def populateCols(self, col):
        w=self.__ws[str(col+str(self.getRows()[0])):str(col+str(self.getRows()[-1]))]
        values=[]
        x=0
        for x in range(len(w)):
            values.append(w[x][0].value)
        return values
    def addPartner(self, partner):  #Unused
        self.__partners.append(partner)

#Unused
'''class PartnerGroups():
    def __init__(self, parentObj, firstChildObj):
        self.__parentTitle=parentObj.getTitle()
        self.__parent=parentObj
        self.__children=[firstChildObj]
    def getTitle(self):
        return self.__parentTitle
    def getParent(self):
        return self.__parent
    def getChildren(self):
        return self.__children
    def addChild(self, child):
        if child not in self.__children:
            self.__children.append(child)
            return True
        else:
            return False'''


#---Creation of Grouping objects---------------------------------------------------------------------------|

#Uses information gained in getStructs to create a Grouping object. Pos and count are used to calculate the range the object spans. 
#Due to 0-indexing, a control boolean for the last group ensures no data is missed.
def createGroup(ws, pos, count, column, last=False):
    list=[]
    title=str(column[pos-1].value)
    z=abs(pos-count)
    if not last:
        while z < pos+1:
            list.append(z)
            z+=1 
    else: 
        while z < pos+2:
            list.append(z)
            z+=1 
    return Grouping(ws, title, list)

#Iterates through the rows of the "A" column of the given spreadsheet. Whenever the current entry doesn't match the last, a new Grouping obj is created.
#Returns a list of Grouping objects that comprise all data within the spreadsheet.
def getStructs(sheet):
    col=sheet['A']
    listOfObj=[]
    count=0
    for x in range(2,len(col)):
        if x!=0:
            prev=col[x-1].value
        else:
            prev=col[x].value        
        if col[x].value!=prev:
            groupObject=createGroup(sheet, x, count, col)
            listOfObj.append(groupObject)
            count=0
        elif x==len(col)-1:
            groupObject=createGroup(sheet, x, count, col, True)
            listOfObj.append(groupObject)
            count=0
        else:
            count+=1
    return listOfObj

#----------------------------------------------------------------------------------------------------------|


#---Detector methods---------------------------------------------------------------------------------------|

#Determines if a variable is of a non-primative datatype. Returns a boolean and, if applicable, the custom datatype as defined in the 'Typedefs' worksheet as a Grouping object. 
def detectNonPrim(groupObj, num):
    dataTypes=['REAL64', 'INT64', 'char', ''] #The only 3 primatives used in the schema. The blank string accounts for 'Same as' entries.
    bCellVal=groupObj.getBCol()[num]
    if not bCellVal in dataTypes:
        subStructs=getStructs(wb['Typedefs'])
        for substruct in subStructs:
            if substruct.getTitle()==bCellVal:
                return True, substruct
    return False, ''

#Does a check for either bracket character, the existence of which denotes an array datatype. Returns a boolean and, if applicable, the value within the brackets
def detectArray(cColVal):
    if '[' in cColVal or ']' in cColVal:
        index=cColVal.rfind('[')
        return True, cColVal[index+1:-1]
    return False,''

#The convention of the NH5RDB schema document is to include the type of vector (and worksheet name) in the name of each dataset. 
#However, in certain "Same as" entries, the sheet name is missing; the convention is inconsistent
#This function checks if the sheet name is present and prepends it if necessary, returning a string that follows convention and can be used.
def detectAndFixSheetName(suspectedTitle):
    sheetTitle=ws.title
    if sheetTitle not in suspectedTitle[:len(sheetTitle)]:
        return (sheetTitle+"_"+suspectedTitle)
    else:
        return (suspectedTitle)

#Does a check for the string "Same as". This convention is accurate as of 1/25/23
def detectSameAs(value):
    return 'Same as' in value

#----------------------------------------------------------------------------------------------------------|


#---Locator methods----------------------------------------------------------------------------------------|

#Given a grouping object that is described with a "Same as", searches master list of all grouping objects in the sheet an returns the one being referred to.
#Should always successfully find.
def locateSameAs(groupObj):
    justTheSame=groupObj.getCCol()[0][8:].replace("/","_") #Removes the phrase "Same as" and fixes the rest of the string to match convention
    found=False
    justTheSame=detectAndFixSheetName(justTheSame)
    for item in structGroupObjList:
        if justTheSame == item.getTitle():
            found=True
            if item not in groupObj.getPartners():
                groupObj.addPartner(item)
            return item
    if not found:
        return None

#Finds all 'family members' of a Grouping object by searching its title within the list of lists generated by findFamilialGroups().
#Returns either the list of family members or a list of just the Grouping object title.
def findPartners(groupObj):
    title=groupObj.getTitle()
    for item in families:
        if title in item:
            return item
    return [title]

#Iterates down the "C" column to find "Same as" entries. When one is found, the dataset being pointed to is found. 
#The running compilation of 'family' lists is searched to see if the dataset already exists within a family. If it is, the current dataset is added to the family.
#If not, a new list is created with the 'parent' and 'child' dataset and added to the list of lists, which is returned.
def findFamilialGroups():
    a=ws["A"]
    c=ws["C"]
    listOfLists=[]
    for x in range (2, len(c)):
        if detectSameAs(c[x].value):
            name=a[x].value
            justTheSame=detectAndFixSheetName(c[x].value[8:].replace("/","_"))
            found=False
            for item in listOfLists:
                if justTheSame in item:
                    found=True
                    item.append(name)
                    break
            if not found:
                listOfLists.append([name, justTheSame])
    return listOfLists

#----------------------------------------------------------------------------------------------------------|


#---String generator methods-------------------------------------------------------------------------------|

#Appends comments to the end of struct member constructor lines. Includes numbering and, if applicable, comments describing the member. 
def appendVarTypeComment(dColVal, num):
    if dColVal != None:
        return "\t"+r'// '+str(num)+"\t"+str(dColVal)
    else:
        return "\t"+r'// '+str(num)

#Finds any datasets that are related to groupObj and writes upper-level case statements
def generateUpperCaseStatement(groupObj):
    partners=findPartners(groupObj)
    upperCases=''
    for item in partners:
        upperCases+= ("case NH5RDB_DATASETS::"+item+":"+"\n")
    upperCases+="{\n"
    return upperCases

#Iterates over the rows of a groupObj and writes case statements for each member
#Detects if dataset is a Same as or if member is an array
def generateLowCaseStatement(groupObj):
    caseCounter=0
    fullCases=''
    originalTitle=groupObj.getTitle()
    if detectSameAs(groupObj.getCCol()[0]):
        groupObj=locateSameAs(groupObj)
    for item in groupObj.getRows():
        isArray, bracketInd=detectArray(groupObj.getCCol()[caseCounter])
        if isArray:
            index=groupObj.getCCol()[caseCounter].rfind('[')
            cColVal=groupObj.getCCol()[caseCounter][:index]
            fullCases+= ("case "+str(caseCounter)+": \t\t{ fVal = (REAL32)"+originalTitle+"[i]."+cColVal+"[sVec.nArrayIndex];"+"\t\tpData->Add((INT32)"+originalTitle+"[i]."+groupObj.getCCol()[0]+", sVec.nColumn, &fVal); } break; "+"\n")
        else:
            fullCases+= ("case "+str(caseCounter)+": \t\t{ fVal = (REAL32)"+originalTitle+"[i]."+groupObj.getCCol()[caseCounter]+";"+"\t\tpData->Add((INT32)"+originalTitle+"[i]."+groupObj.getCCol()[0]+", sVec.nColumn, &fVal); } break; "+"\n")
        caseCounter+=1
    return fullCases

#----------------------------------------------------------------------------------------------------------|


#---Writer methods-----------------------------------------------------------------------------------------|

#Entry point to defining the structs. Creates the header and closer strings
def makeStruct(groupObj):
    f.write ("struct "+ str(groupObj.getTitle())+'\n')
    f.write ("{"+'\n')
    makeStructConstructorBody(groupObj)
    f.write ("};\n\n"+'\n')

#Takes a grouping obj and writes constructors for the dataset's members using data stored in the object.
#If dataset is described with "Same as", recurs with dataset described. If member is a custom type, recurs with datatype described. Both are stored in Grouping objs.
#Optional params keep information consitant across recursions, allow for accurate numbering and array sizing. 
#Returns a counter variable to keep numbering consistant across recursions
def makeStructConstructorBody(groupObj, control=0, areArrays=(False,'')):
    if detectSameAs(groupObj.getCCol()[0]):
        parent=locateSameAs(groupObj)
        f.write ("\t"+r'// '+str(groupObj.getCCol()[0]).replace("/","_")+"\n") #Adds 'Same as...' comment, useful for debugging
        makeStructConstructorBody(parent) 
    else:
        for num in range(len(groupObj.getRows())): #Iterates through rows of the dataset
                nonPrimAndObj=detectNonPrim(groupObj, num)
                isArray=detectArray(groupObj.getCCol()[num])
                if nonPrimAndObj[0]:
                    control=makeStructConstructorBody(nonPrimAndObj[1],control, isArray)-1
                else:
                    #The difference between the below print statements is the top declares the constructed member as an array. Are otherwise identical
                    if areArrays[0]:
                        f.write ("\t"+str(groupObj.getBCol()[num])+"\t"+str(groupObj.getCCol()[num])+r'['+str(areArrays[1])+r']'+";"+appendVarTypeComment(groupObj.getDCol()[num], control)+'\n')
                    else:
                        f.write ("\t"+str(groupObj.getBCol()[num])+"\t"+str(groupObj.getCCol()[num])+";"+appendVarTypeComment(groupObj.getDCol()[num], control)+'\n')
                control+=1
    return control

#Inserts both upper and lower level case statements into a static logic tree
#Writes output to text file
def inputIntoTree(groupObj):
    upperCases=generateUpperCaseStatement(groupObj)
    cases=generateLowCaseStatement(groupObj)
    title=groupObj.getTitle()
    tree='''if (read_all)
			{
				if (within_single_domain)
				{
					for (i = 0; i < num_elements_to_read; i++)
					{
						for (const auto& sVec : (*m_vectors_by_dataset[dataset]))
						{
							switch (sVec.nFieldIndex)
							{
                                '''+cases+'''
							}
						}
					}
				}
				else
				{
					for (i = 0; i < num_elements_to_read; i++)
					{
						for (const auto& sVec : (*m_vectors_by_dataset[dataset]))
						{
							if (sVec.nDomainID == '''+title+'''[i].DOMAIN_ID)
							{
								switch (sVec.nFieldIndex)
								{
								    '''+cases+'''
								}
							}
						}
					}

				}
			}
			else
			{
				if (within_single_domain)
				{
					for (i = 0; i < num_elements_to_read; i++)
					{
						if (pID->IsAdded((INT32)'''+title+'''[i].EID))
						{
							for (const auto& sVec : (*m_vectors_by_dataset[dataset]))
							{
								switch (sVec.nFieldIndex)
								{
								    '''+cases+'''
								}
							}
						}
					}

				}
				else
				{
					for (i = 0; i < num_elements_to_read; i++)
					{
						if (pID->IsAdded((INT32)'''+title+'''[i].EID))
						{
							for (const auto& sVec : (*m_vectors_by_dataset[dataset]))
							{
								if (sVec.nDomainID == '''+title+'''[i].DOMAIN_ID)
								{
									switch (sVec.nFieldIndex)
									{
									    '''+cases+'''
									}
								}
							}
						}
					}
				}
			}
        '''
    f.write (upperCases+tree+"}\n\n\n")

#Writes pointers(?) that are included at the bottom of NH5RDBRead.h. Likely will need custom formatting.
def writePointers(groupObjList):
    for item in groupObjList:
        f.write ("\tstruct "+item.getTitle()+"\t\t\t*"+item.getTitle()+";\n")



#Unused
def getPartners(listOfObj):
    return ''
    partnerObjs=[]
    partnerObjTitles=[]
    for item in listOfObj:
        if detectSameAs(item.getCCol()[0]):
            sameAs=locateSameAs(item)
            if sameAs.getTitle() not in partnerObjTitles:
                newPartnerObj=PartnerGroups(sameAs, item)
                partnerObjs.append(newPartnerObj)
                partnerObjTitles.append(newPartnerObj.getTitle())
            else:
                for obj in partnerObjs:
                    if obj.getTitle()==sameAs.getTitle():
                        rc=obj.addChild(sameAs)
                        break
    return partnerObjs, partnerObjTitles



#Main method. Contains controls for getting different outputs. 
#The text files are closed after each use to reuse the variable 'f', avoiding the need to pass an arg through every method in a process.
if __name__=="__main__":

    wb=openpyxl.load_workbook("Convention_To_Code.xlsx") #Workbook conatining MSC Nastan hdf5 schema data. Must specify full path if not in same folder as script.
    ws=wb.active #Set the worksheet to whichever is active. Will be overwritten.

    #----------------------------------------------------------------------------------------------------------|
    #Creating struct implementations for datasets (NH5RDBRead.h)
    f=open('STRUCTDEFS.txt','w')

    ws=wb["ELEMENTAL"]              #Setting global variables to elemental values
    structGroupObjList=getStructs(ws)   #
    for item in structGroupObjList:
        makeStruct(item)

    ws=wb["NODAL"]
    structGroupObjList=getStructs(ws)
    for item in structGroupObjList:
        makeStruct(item)

    f.close()
    #----------------------------------------------------------------------------------------------------------|


    #----------------------------------------------------------------------------------------------------------|
    #Creating case statements for datasets (NH5RDBRead.cpp)
    f=open('CASES.txt', 'w')
    f.write("switch (dataset)\n{\n")

    #ws and structGroupObjList are still in NODAL mode from previous operation
    families=findFamilialGroups()
    for item in structGroupObjList:
        if not detectSameAs(item.getCCol()[0]):
            inputIntoTree(item)

    ws=wb["ELEMENTAL"]              #Setting global variables to elemental values
    structGroupObjList=getStructs(ws)   #
    families=findFamilialGroups()
    for item in structGroupObjList:
        if not detectSameAs(item.getCCol()[0]):
            inputIntoTree(item)

    f.write("}")
    f.close()
    #----------------------------------------------------------------------------------------------------------|


    #----------------------------------------------------------------------------------------------------------|
    #Creating pointers(?) for each struct in the spreadsheet (NH5RDBRead.h)
    f=open("STRUCTPOINTERS.txt", 'w')

    #ws and structGroupObjList are still in ELEMENTAL mode from previous operation
    writePointers(structGroupObjList)
    
    ws=wb["NODAL"]  #Setting to NODAL mode
    structGroupObjList=getStructs(ws) #
    writePointers(structGroupObjList)

    f.close()
    #----------------------------------------------------------------------------------------------------------|



    print ("Done")